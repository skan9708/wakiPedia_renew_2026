import openpyxl
from django.core.management.base import BaseCommand
from django.db.models import Q

from wine.models import Wine, WineReview, Tag


def parse_rating(raw) -> int:
    """평점(0~5 소수)을 DB 단위(0~50, 5배수)로 변환"""
    try:
        val = float(raw) * 10
        val = round(val / 5) * 5
        return max(0, min(50, int(val)))
    except (TypeError, ValueError):
        return 30  # 기본 3.0점


def find_wine(eng_name: str, kor_name: str):
    """영문명 또는 한글명으로 와인 검색 (정확 → 부분일치 순)"""
    # 1. 두 이름 모두 정확 일치
    if eng_name and kor_name:
        w = Wine.objects.filter(eng_name__iexact=eng_name, kor_name__iexact=kor_name).first()
        if w:
            return w

    # 2. 영문명 정확 일치
    if eng_name:
        w = Wine.objects.filter(eng_name__iexact=eng_name).first()
        if w:
            return w

    # 3. 한글명 정확 일치
    if kor_name:
        w = Wine.objects.filter(kor_name__iexact=kor_name).first()
        if w:
            return w

    # 4. 영문명 부분 일치 (앞 20자로 축약 검색)
    if eng_name and len(eng_name) > 4:
        prefix = eng_name[:20].strip()
        w = Wine.objects.filter(eng_name__icontains=prefix).first()
        if w:
            return w

    # 5. 한글명 부분 일치
    if kor_name and len(kor_name) > 1:
        prefix = kor_name[:10].strip()
        w = Wine.objects.filter(kor_name__icontains=prefix).first()
        if w:
            return w

    return None


class Command(BaseCommand):
    help = '엑셀 파일에서 [이유]/[키워드]를 읽어 일치하는 와인에 리뷰를 등록합니다.'

    def add_arguments(self, parser):
        parser.add_argument('xlsx_file', type=str)
        parser.add_argument('--user-id', type=int, default=1, help='리뷰 작성자 user id (기본: 1)')
        parser.add_argument('--dry-run', action='store_true', help='DB에 저장하지 않고 결과만 출력')

    def handle(self, *args, **options):
        from user.models import User
        try:
            author = User.objects.get(id=options['user_id'])
        except User.DoesNotExist:
            self.stdout.write(self.style.ERROR(f'user id {options["user_id"]} 없음'))
            return

        wb = openpyxl.load_workbook(options['xlsx_file'], data_only=True)

        # 시트1(원본) 찾기 (이름 인코딩 무관하게 첫 번째 시트 사용)
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]

        # row1=메타, row2=헤더, row3~=데이터
        rows = list(ws.iter_rows(min_row=3, values_only=True))
        wb.close()
        self.stdout.write(f'시트: {sheet_name}  데이터 행: {len(rows)}')

        created = skipped_no_wine = skipped_no_content = skipped_dup = 0
        matched_wines = []

        for row in rows:
            if not row or all(v is None for v in row):
                continue

            # 컬럼 매핑: 번호, 영문명, 한글명, 평점, 이유, 키워드, ...
            eng_name = str(row[1] or '').strip()
            kor_name = str(row[2] or '').strip()
            raw_rating = row[3]
            reason   = str(row[4] or '').strip()   # 이유
            keywords = str(row[5] or '').strip()   # 키워드

            if not eng_name and not kor_name:
                continue

            # 리뷰 텍스트 결정
            text = reason or keywords
            if not text:
                skipped_no_content += 1
                continue

            wine = find_wine(eng_name, kor_name)
            if not wine:
                skipped_no_wine += 1
                continue

            # 중복 확인 (같은 user + wine + text 앞 30자)
            if WineReview.objects.filter(wine=wine, user=author, text__startswith=text[:30]).exists():
                skipped_dup += 1
                continue

            matched_wines.append((wine.kor_name, text[:40]))

            if options['dry_run']:
                continue

            rating = parse_rating(raw_rating)

            review = WineReview.objects.create(
                wine=wine,
                user=author,
                text=text,
                rating=rating,
                body=2,
                acidity=2,
                sweetness=2,
                tannin=2,
                is_public=True,
            )

            # 키워드 → 태그
            if keywords:
                for kw in keywords.split(','):
                    kw = kw.strip()
                    if kw:
                        tag, _ = Tag.objects.get_or_create(name=kw, defaults={'category': 'fruity'})
                        review.tags.add(tag)

            created += 1

        self.stdout.write(self.style.SUCCESS(f'\n완료: {created}개 리뷰 등록'))
        self.stdout.write(f'매칭 성공: {len(matched_wines)} / 와인 미매칭: {skipped_no_wine} / 내용없음: {skipped_no_content} / 중복: {skipped_dup}')

        if options['dry_run'] and matched_wines:
            self.stdout.write('\n[dry-run] 매칭된 와인 샘플:')
            for name, txt in matched_wines[:20]:
                self.stdout.write(f'  {name} → {txt}')
